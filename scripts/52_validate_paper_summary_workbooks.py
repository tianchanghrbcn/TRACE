#!/usr/bin/env python3
"""Validate generated paper summary workbooks against archived Mode A summaries."""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

TRACE_PROJECT_ROOT = Path(
    os.environ.get("TRACE_PROJECT_ROOT", Path(__file__).resolve().parents[1])
).resolve()

if str(TRACE_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(TRACE_PROJECT_ROOT))

from openpyxl import load_workbook

from src.paper_replay.workbook_utils import sha256_file, write_json


DATASETS = ["beers", "flights", "hospital", "rayyan"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate generated paper summary workbooks.")
    parser.add_argument(
        "--manifest",
        type=Path,
        default=Path("analysis/paper_generated/generated_summary_manifest.json"),
    )
    parser.add_argument(
        "--archived-summary-root",
        type=Path,
        default=Path("analysis/paper_exact/analysis_summaries/results/analysis_results"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("analysis/paper_generated/generated_summary_validation_report.json"),
    )
    return parser.parse_args()


def workbook_shape(path: Path) -> dict:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        sheets[name] = {
            "max_row": ws.max_row,
            "max_column": ws.max_column,
        }
    wb.close()
    return sheets


def find_archived_summary(root: Path, dataset: str) -> Path | None:
    candidates = sorted(root.rglob(f"{dataset}_summary.xlsx"))
    return candidates[0] if candidates else None


def main() -> None:
    args = parse_args()

    if not args.manifest.exists():
        raise FileNotFoundError(f"Generated summary manifest not found: {args.manifest}")

    manifest = json.loads(args.manifest.read_text(encoding="utf-8-sig"))
    generated_by_dataset = {
        item["dataset"]: Path(item["output_xlsx"])
        for item in manifest.get("datasets", [])
    }

    rows = []
    failures = []

    for dataset in DATASETS:
        generated = generated_by_dataset.get(dataset)
        archived = find_archived_summary(args.archived_summary_root, dataset)

        row = {
            "dataset": dataset,
            "generated_path": str(generated) if generated else "",
            "generated_exists": bool(generated and generated.exists()),
            "archived_path": str(archived) if archived else "",
            "archived_exists": bool(archived and archived.exists()),
            "generated_sha256": "",
            "archived_sha256": "",
            "generated_sheets": {},
            "archived_sheets": {},
        }

        if not row["generated_exists"]:
            failures.append(f"{dataset}: generated summary workbook missing.")
        else:
            row["generated_sha256"] = sha256_file(generated)
            row["generated_sheets"] = workbook_shape(generated)

            # Required v0 sheets generated from archived analysis CSVs.
            for required in ["overview", "cleaning", "cluster"]:
                if required not in row["generated_sheets"]:
                    failures.append(f"{dataset}: generated workbook missing required sheet `{required}`.")

        if not row["archived_exists"]:
            failures.append(f"{dataset}: archived paper summary workbook missing.")
        else:
            row["archived_sha256"] = sha256_file(archived)
            row["archived_sheets"] = workbook_shape(archived)

        rows.append(row)

    index_path = Path(manifest.get("index_workbook", {}).get("output_xlsx", ""))
    if not index_path.exists():
        failures.append("paper_summary_index.xlsx missing.")

    report = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": "PASS" if not failures else "FAIL",
        "manifest": str(args.manifest),
        "archived_summary_root": str(args.archived_summary_root),
        "datasets": rows,
        "failures": failures,
        "note": (
            "This validates generated summary workbooks from archived analysis CSVs. "
            "It does not yet claim byte-identical reproduction of every paper table."
        ),
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    write_json(args.output, report)

    md_path = args.output.with_suffix(".md")
    lines = [
        "# Generated Paper Summary Workbook Validation",
        "",
        f"- Generated at UTC: {report['generated_at_utc']}",
        f"- Status: {report['status']}",
        "",
        "## Dataset workbooks",
        "",
        "| Dataset | Generated exists | Archived exists | Generated sheets | Archived sheets |",
        "|---|---:|---:|---|---|",
    ]

    for row in rows:
        lines.append(
            f"| {row['dataset']} | {row['generated_exists']} | {row['archived_exists']} | "
            f"{', '.join(row['generated_sheets'].keys())} | {', '.join(row['archived_sheets'].keys())} |"
        )

    lines += [
        "",
        "## Failures",
        "",
    ]

    if failures:
        for failure in failures:
            lines.append(f"- {failure}")
    else:
        lines.append("No failures.")

    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(json.dumps(report, indent=2, ensure_ascii=False))
    print(f"[TRACE] Validation report written to: {args.output}")
    print(f"[TRACE] Validation markdown written to: {md_path}")
    print(f"[TRACE] Generated summary workbook status: {report['status']}")

    raise SystemExit(0 if report["status"] == "PASS" else 1)


if __name__ == "__main__":
    main()

