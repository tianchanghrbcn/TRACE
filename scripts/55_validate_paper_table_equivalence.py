#!/usr/bin/env python3
"""Validate equivalence between generated paper-table outputs and archived references.

This script compares paper-table outputs collected by Stage 3R.4.1 against
archived paper-exact references. It supports:

- exact SHA-256 match;
- semantic CSV comparison with numeric tolerance;
- semantic XLSX comparison with numeric tolerance;
- semantic JSON comparison;
- normalized TXT comparison.

The validator intentionally distinguishes hard failures from warnings.
A generated fallback without archived reference is reported as WARN rather
than FAIL when all other comparable outputs are valid.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


COMPARABLE_EXTENSIONS = {".csv", ".xlsx", ".xls", ".json", ".txt"}
DEFAULT_ARCHIVE_ROOTS = [
    Path("analysis/paper_exact"),
    Path("artifacts/paper_exact"),
]
NUMERIC_TOLERANCE = 1.0e-6


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate paper-table output equivalence.")
    parser.add_argument(
        "--table-validation-report",
        type=Path,
        default=Path("analysis/paper_generated/paper_tables/paper_table_validation_report.json"),
    )
    parser.add_argument(
        "--run-manifest",
        type=Path,
        default=Path("analysis/paper_generated/paper_tables/paper_table_script_run_manifest.json"),
    )
    parser.add_argument(
        "--archive-roots",
        nargs="+",
        type=Path,
        default=DEFAULT_ARCHIVE_ROOTS,
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("analysis/paper_generated/paper_tables/table_equivalence_report.json"),
    )
    parser.add_argument(
        "--diff-dir",
        type=Path,
        default=Path("analysis/paper_generated/paper_tables/table_diffs"),
    )
    parser.add_argument("--tolerance", type=float, default=NUMERIC_TOLERANCE)
    return parser.parse_args()


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Required JSON not found: {path}")
    return json.loads(path.read_text(encoding="utf-8-sig", errors="replace"))


def write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def sha256_file(path: Path) -> str:
    import hashlib

    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def try_float(value: Any) -> float | None:
    value = normalize_cell(value)
    if value == "":
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    try:
        text = str(value).strip().replace("%", "")
        if text == "":
            return None
        return float(text)
    except Exception:
        return None


def values_equal(a: Any, b: Any, tolerance: float) -> bool:
    fa = try_float(a)
    fb = try_float(b)

    if fa is not None and fb is not None:
        return abs(fa - fb) <= tolerance

    return str(normalize_cell(a)) == str(normalize_cell(b))


def compare_csv(a: Path, b: Path, tolerance: float) -> tuple[bool, dict[str, Any]]:
    da = pd.read_csv(a, encoding="utf-8-sig")
    db = pd.read_csv(b, encoding="utf-8-sig")

    da.columns = [str(c).strip() for c in da.columns]
    db.columns = [str(c).strip() for c in db.columns]

    detail: dict[str, Any] = {
        "kind": "csv",
        "shape_a": list(da.shape),
        "shape_b": list(db.shape),
        "column_match": sorted(da.columns) == sorted(db.columns),
        "mismatch_count": 0,
        "sample_mismatches": [],
    }

    if da.shape[0] != db.shape[0]:
        detail["row_mismatch"] = True
        return False, detail

    if sorted(da.columns) != sorted(db.columns):
        return False, detail

    db = db[da.columns]

    mismatch_count = 0
    samples = []

    for col in da.columns:
        for idx, (va, vb) in enumerate(zip(da[col].tolist(), db[col].tolist())):
            if not values_equal(va, vb, tolerance):
                mismatch_count += 1
                if len(samples) < 20:
                    samples.append(
                        {
                            "row": idx,
                            "column": col,
                            "generated": str(va),
                            "archived": str(vb),
                        }
                    )

    detail["mismatch_count"] = mismatch_count
    detail["sample_mismatches"] = samples
    return mismatch_count == 0, detail


def workbook_values(path: Path) -> dict[str, list[list[Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([normalize_cell(value) for value in row])
        data[sheet_name] = rows

    wb.close()
    return data


def compare_xlsx(a: Path, b: Path, tolerance: float) -> tuple[bool, dict[str, Any]]:
    wa = workbook_values(a)
    wb = workbook_values(b)

    sheets_a = list(wa.keys())
    sheets_b = list(wb.keys())

    detail: dict[str, Any] = {
        "kind": "xlsx",
        "sheets_a": sheets_a,
        "sheets_b": sheets_b,
        "sheet_match": sheets_a == sheets_b,
        "mismatch_count": 0,
        "sample_mismatches": [],
    }

    # Strict path: same sheet names.
    if sheets_a == sheets_b:
        sheet_pairs = [(name, name) for name in sheets_a]
    # Fallback: one sheet on each side, compare them despite name difference.
    elif len(sheets_a) == 1 and len(sheets_b) == 1:
        sheet_pairs = [(sheets_a[0], sheets_b[0])]
        detail["sheet_match"] = "single_sheet_name_fallback"
    else:
        return False, detail

    mismatch_count = 0
    samples = []

    for sa, sb in sheet_pairs:
        rows_a = wa[sa]
        rows_b = wb[sb]

        if len(rows_a) != len(rows_b):
            detail["row_count_mismatch"] = {
                "sheet_a": sa,
                "sheet_b": sb,
                "rows_a": len(rows_a),
                "rows_b": len(rows_b),
            }
            return False, detail

        for i, (ra, rb) in enumerate(zip(rows_a, rows_b)):
            width = max(len(ra), len(rb))
            aa = list(ra) + [""] * (width - len(ra))
            bb = list(rb) + [""] * (width - len(rb))

            for j, (va, vb) in enumerate(zip(aa, bb)):
                if not values_equal(va, vb, tolerance):
                    mismatch_count += 1
                    if len(samples) < 20:
                        samples.append(
                            {
                                "sheet_generated": sa,
                                "sheet_archived": sb,
                                "row": i + 1,
                                "column": j + 1,
                                "generated": str(va),
                                "archived": str(vb),
                            }
                        )

    detail["mismatch_count"] = mismatch_count
    detail["sample_mismatches"] = samples
    return mismatch_count == 0, detail


def normalize_json_obj(obj: Any) -> Any:
    if isinstance(obj, dict):
        return {str(k): normalize_json_obj(v) for k, v in sorted(obj.items())}
    if isinstance(obj, list):
        return [normalize_json_obj(v) for v in obj]
    if isinstance(obj, float):
        return round(obj, 10)
    return obj


def compare_json(a: Path, b: Path, tolerance: float) -> tuple[bool, dict[str, Any]]:
    try:
        oa = normalize_json_obj(json.loads(a.read_text(encoding="utf-8-sig")))
        ob = normalize_json_obj(json.loads(b.read_text(encoding="utf-8-sig")))
    except Exception as exc:
        return False, {"kind": "json", "error": repr(exc)}

    return oa == ob, {"kind": "json", "equal": oa == ob}


def compare_text(a: Path, b: Path, tolerance: float) -> tuple[bool, dict[str, Any]]:
    ta = a.read_text(encoding="utf-8-sig", errors="replace").replace("\r\n", "\n").strip()
    tb = b.read_text(encoding="utf-8-sig", errors="replace").replace("\r\n", "\n").strip()

    return ta == tb, {
        "kind": "txt",
        "len_a": len(ta),
        "len_b": len(tb),
        "equal": ta == tb,
    }


def compare_files(a: Path, b: Path, tolerance: float) -> tuple[str, dict[str, Any]]:
    if not a.exists() or not b.exists():
        return "FAIL", {"reason": "missing_file"}

    if a.suffix.lower() != b.suffix.lower():
        return "FAIL", {
            "reason": "extension_mismatch",
            "generated_ext": a.suffix.lower(),
            "archived_ext": b.suffix.lower(),
        }

    digest_a = sha256_file(a)
    digest_b = sha256_file(b)

    if digest_a == digest_b:
        return "EXACT", {
            "reason": "sha256_match",
            "sha256": digest_a,
        }

    ext = a.suffix.lower()
    if ext == ".csv":
        equal, detail = compare_csv(a, b, tolerance)
    elif ext in {".xlsx", ".xls"}:
        equal, detail = compare_xlsx(a, b, tolerance)
    elif ext == ".json":
        equal, detail = compare_json(a, b, tolerance)
    elif ext == ".txt":
        equal, detail = compare_text(a, b, tolerance)
    else:
        return "WARN", {"reason": f"unsupported_extension:{ext}"}

    detail["sha256_generated"] = digest_a
    detail["sha256_archived"] = digest_b

    return ("SEMANTIC" if equal else "FAIL"), detail


def extract_path_from_generated_match(text: str) -> Path | None:
    # generated_matches are formatted like:
    #   relative_path E:\TRACE\analysis\...\output.csv
    parts = text.split()
    for part in reversed(parts):
        candidate = Path(part)
        if candidate.exists():
            return candidate
    return None


def collect_generated_paths(report: dict[str, Any], run_manifest: dict[str, Any]) -> list[Path]:
    paths: list[Path] = []

    for row in report.get("expected_output_checks", []):
        for path_str in row.get("copied_available_outputs", []):
            path = Path(path_str)
            if path.exists() and path.suffix.lower() in COMPARABLE_EXTENSIONS:
                paths.append(path)

        for match in row.get("generated_matches", []):
            path = extract_path_from_generated_match(match)
            if path and path.exists() and path.suffix.lower() in COMPARABLE_EXTENSIONS:
                paths.append(path)

    for row in run_manifest.get("collected_outputs", []):
        path = Path(row.get("output_path", ""))
        if path.exists() and path.suffix.lower() in COMPARABLE_EXTENSIONS:
            paths.append(path)

    # Deduplicate while preserving deterministic order.
    seen = set()
    out = []
    for path in sorted(paths, key=lambda p: str(p).lower()):
        key = str(path.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(path)

    return out


def archive_candidates(file_name: str, archive_roots: list[Path]) -> list[Path]:
    candidates = []

    for root in archive_roots:
        if not root.exists():
            continue

        for path in root.rglob(file_name):
            if not path.is_file():
                continue

            lower = str(path).lower()
            if "paper_generated" in lower:
                continue

            candidates.append(path)

    seen = set()
    out = []
    for path in sorted(candidates, key=lambda p: str(p).lower()):
        key = str(path.resolve()).lower()
        if key not in seen:
            seen.add(key)
            out.append(path)

    return out


def best_match(generated: Path, candidates: list[Path], tolerance: float) -> dict[str, Any]:
    if not candidates:
        return {
            "status": "WARN_NO_REFERENCE",
            "best_reference": "",
            "detail": {"reason": "no archived reference with same filename"},
        }

    attempts = []

    for candidate in candidates:
        status, detail = compare_files(generated, candidate, tolerance)
        attempts.append(
            {
                "reference": str(candidate),
                "status": status,
                "detail": detail,
            }
        )

        if status in {"EXACT", "SEMANTIC"}:
            return {
                "status": status,
                "best_reference": str(candidate),
                "detail": detail,
                "candidate_count": len(candidates),
            }

    return {
        "status": "FAIL",
        "best_reference": attempts[0]["reference"] if attempts else "",
        "detail": {
            "reason": "no equivalent archived reference",
            "attempts": attempts[:5],
        },
        "candidate_count": len(candidates),
    }


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    columns = [
        "generated_path",
        "file_name",
        "status",
        "best_reference",
        "candidate_count",
    ]

    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, report: dict[str, Any]) -> None:
    lines = [
        "# Paper Table Equivalence Validation",
        "",
        f"- Generated at UTC: {report['generated_at_utc']}",
        f"- Status: {report['status']}",
        f"- Compared files: {report['compared_file_count']}",
        "",
        "## Status counts",
        "",
        "| Status | Count |",
        "|---|---:|",
    ]

    for key, value in sorted(report["status_counts"].items()):
        lines.append(f"| {key} | {value} |")

    lines += [
        "",
        "## Comparisons",
        "",
        "| Status | Generated file | Best reference |",
        "|---|---|---|",
    ]

    for row in report["comparisons"]:
        lines.append(
            f"| {row['status']} | {row['generated_path']} | {row.get('best_reference', '')} |"
        )

    lines += [
        "",
        "## Failures",
        "",
    ]

    if report["failures"]:
        for failure in report["failures"]:
            lines.append(f"- {failure}")
    else:
        lines.append("No hard failures.")

    lines += [
        "",
        "## Scope note",
        "",
        "This validates table-output equivalence between generated/available Mode A outputs and archived references.",
        "It does not yet validate final paper figure equivalence or all narrative claims in the paper.",
    ]

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    args = parse_args()

    validation_report = read_json(args.table_validation_report)
    run_manifest = read_json(args.run_manifest)

    generated_paths = collect_generated_paths(validation_report, run_manifest)

    comparisons = []
    failures = []

    for generated in generated_paths:
        candidates = archive_candidates(generated.name, args.archive_roots)
        result = best_match(generated, candidates, args.tolerance)
        row = {
            "generated_path": str(generated),
            "file_name": generated.name,
            "status": result["status"],
            "best_reference": result.get("best_reference", ""),
            "candidate_count": result.get("candidate_count", 0),
            "detail": result.get("detail", {}),
        }
        comparisons.append(row)

        if row["status"] == "FAIL":
            failures.append(
                f"{generated} has no equivalent archived reference."
            )

    counts = Counter(row["status"] for row in comparisons)

    status = "PASS"
    if failures:
        status = "FAIL"
    elif counts.get("WARN_NO_REFERENCE", 0) or counts.get("WARN", 0):
        status = "PASS_WITH_WARNINGS"

    report = {
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": status,
        "table_validation_report": str(args.table_validation_report),
        "run_manifest": str(args.run_manifest),
        "archive_roots": [str(root) for root in args.archive_roots],
        "tolerance": args.tolerance,
        "compared_file_count": len(comparisons),
        "status_counts": dict(counts),
        "comparisons": comparisons,
        "failures": failures,
        "note": (
            "PASS means generated/available comparable paper-table outputs match archived references. "
            "PASS_WITH_WARNINGS means no hard mismatches were found, but some outputs had no archived reference."
        ),
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    write_json(args.output, report)
    write_csv(args.output.with_suffix(".csv"), comparisons)
    write_markdown(args.output.with_suffix(".md"), report)

    print(json.dumps(
        {
            "status": report["status"],
            "compared_file_count": report["compared_file_count"],
            "status_counts": report["status_counts"],
            "failure_count": len(failures),
            "output": str(args.output),
        },
        indent=2,
        ensure_ascii=False,
    ))
    print(f"[TRACE] Paper table equivalence report written to: {args.output}")
    print(f"[TRACE] Paper table equivalence status: {report['status']}")

    raise SystemExit(0 if status in {"PASS", "PASS_WITH_WARNINGS"} else 1)


if __name__ == "__main__":
    main()

